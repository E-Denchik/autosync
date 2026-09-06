import pytest

from app.extensions import db
from app.models import (
    Contract,
    ConfidenceLevel,
    ContractPart,
    DocumentProcessingStatus,
    PartMatch,
    RepairOrder,
    RepairOrderStatus,
    ReviewStatus,
)


@pytest.fixture
def repair_order_with_matches(app):
    with app.app_context():
        contract = Contract(
            original_filename="contract.xlsx",
            storage_path="/tmp/contract.xlsx",
            status=DocumentProcessingStatus.PARSED,
        )
        db.session.add(contract)
        db.session.flush()

        order = RepairOrder(
            contract_id=contract.id,
            original_filename="order.xlsx",
            storage_path="/tmp/order.xlsx",
            status=RepairOrderStatus.NEEDS_REVIEW,
            parsed_lines=[{"article": "ABC-1", "name": "Диск", "qty": 1, "price": 1000.0}],
        )
        db.session.add(order)
        db.session.flush()

        m1 = PartMatch(
            repair_order_id=order.id,
            contract_article="ABC-1",
            contract_name="Тормозной диск",
            matched_article="ABC-1",
            matched_name="Диск тормозной",
            matched_price=1000.0,
            confidence_level=ConfidenceLevel.EXACT,
            confidence_score=1.0,
            review_status=ReviewStatus.PENDING,
        )
        m2 = PartMatch(
            repair_order_id=order.id,
            contract_article="XYZ-9",
            contract_name="Фильтр",
            matched_article=None,
            matched_name=None,
            confidence_level=ConfidenceLevel.LLM_GUESS,
            confidence_score=0.4,
            review_status=ReviewStatus.PENDING,
        )
        db.session.add_all([m1, m2])
        db.session.commit()
        return {"order_id": order.id, "match_ids": [m1.id, m2.id]}


def test_list_matches(client, admin_headers, repair_order_with_matches):
    order_id = repair_order_with_matches["order_id"]
    resp = client.get(f"/api/repair-orders/matching/{order_id}", headers=admin_headers)
    assert resp.status_code == 200
    body = resp.get_json()
    assert len(body) == 2
    assert {m["confidence_level"] for m in body} == {"exact", "llm_guess"}


def test_list_exposes_llm_error_so_reviewer_can_tell_it_apart_from_genuine_no_match(client, admin_headers, app):
    """Регрессия по прозрачности: раньше "ИИ была недоступна" и "ИИ честно
    не нашла совпадение" выглядели для проверяющего одинаково."""
    with app.app_context():
        contract = Contract(original_filename="c.xlsx", storage_path="/tmp/c.xlsx", status=DocumentProcessingStatus.PARSED)
        db.session.add(contract)
        db.session.flush()
        order = RepairOrder(
            contract_id=contract.id, original_filename="o.xlsx", storage_path="/tmp/o.xlsx", status=RepairOrderStatus.NEEDS_REVIEW
        )
        db.session.add(order)
        db.session.flush()
        db.session.add(
            PartMatch(
                repair_order_id=order.id,
                contract_name="Деталь",
                confidence_level=ConfidenceLevel.LLM_GUESS,
                confidence_score=0.0,
                review_status=ReviewStatus.PENDING,
                raw_match_data={"source": "llm_error", "error": "llm-service недоступен: Connection refused"},
            )
        )
        db.session.commit()
        order_id = order.id

    body = client.get(f"/api/repair-orders/matching/{order_id}", headers=admin_headers).get_json()
    assert body[0]["llm_error"] == "llm-service недоступен: Connection refused"


def _make_bare_repair_order(app) -> int:
    contract = Contract(original_filename="c.xlsx", storage_path="/tmp/c.xlsx", status=DocumentProcessingStatus.PARSED)
    db.session.add(contract)
    db.session.flush()
    order = RepairOrder(
        contract_id=contract.id, original_filename="o.xlsx", storage_path="/tmp/o.xlsx", status=RepairOrderStatus.NEEDS_REVIEW
    )
    db.session.add(order)
    db.session.commit()
    return order.id


@pytest.mark.parametrize(
    "confidence_level,matched_name,raw_match_data,expected_category",
    [
        (ConfidenceLevel.EXACT, "Деталь", {"source": "exact_article_match"}, "exact"),
        (ConfidenceLevel.CROSS_REF, "Деталь", {"source": "parts_supplier_cross_reference"}, "cross_ref"),
        (ConfidenceLevel.LLM_GUESS, None, {"source": "llm_error", "error": "x"}, "llm_error"),
        (ConfidenceLevel.LLM_GUESS, None, {"source": "no_match_found"}, "no_match"),
        (ConfidenceLevel.LLM_GUESS, "Деталь", {"source": "llm_fallback"}, "llm_guess"),
    ],
)
def test_match_category_classification(
    client, admin_headers, app, confidence_level, matched_name, raw_match_data, expected_category
):
    with app.app_context():
        order_id = _make_bare_repair_order(app)
        db.session.add(
            PartMatch(
                repair_order_id=order_id,
                contract_name="Деталь",
                matched_name=matched_name,
                confidence_level=confidence_level,
                confidence_score=1.0 if confidence_level != ConfidenceLevel.LLM_GUESS else 0.0,
                review_status=ReviewStatus.PENDING,
                raw_match_data=raw_match_data,
            )
        )
        db.session.commit()

    body = client.get(f"/api/repair-orders/matching/{order_id}", headers=admin_headers).get_json()
    assert body[0]["match_category"] == expected_category


@pytest.mark.parametrize(
    "confidence_level,confidence_score,review_status,manually_edited,raw_match_data,expected_is_verified",
    [
        (ConfidenceLevel.EXACT, 1.0, ReviewStatus.PENDING, False, {"source": "exact_article_match"}, True),
        (ConfidenceLevel.CROSS_REF, 0.9, ReviewStatus.PENDING, False, {"source": "parts_supplier_cross_reference"}, True),
        (ConfidenceLevel.LLM_GUESS, 0.9, ReviewStatus.PENDING, False, {"source": "llm_fallback"}, True),
        (ConfidenceLevel.LLM_GUESS, 0.4, ReviewStatus.PENDING, False, {"source": "llm_fallback"}, False),
        (ConfidenceLevel.LLM_GUESS, 0.4, ReviewStatus.PENDING, True, {"source": "llm_fallback"}, True),
    ],
)
def test_is_verified_classification(
    client,
    admin_headers,
    app,
    confidence_level,
    confidence_score,
    review_status,
    manually_edited,
    raw_match_data,
    expected_is_verified,
):
    with app.app_context():
        order_id = _make_bare_repair_order(app)
        db.session.add(
            PartMatch(
                repair_order_id=order_id,
                contract_name="Деталь",
                matched_name="Деталь найденная",
                confidence_level=confidence_level,
                confidence_score=confidence_score,
                review_status=review_status,
                manually_edited=manually_edited,
                raw_match_data=raw_match_data,
            )
        )
        db.session.commit()

    body = client.get(f"/api/repair-orders/matching/{order_id}", headers=admin_headers).get_json()
    assert body[0]["is_verified"] is expected_is_verified


def test_is_verified_false_when_approved_but_no_value_set(client, admin_headers, app):
    """Регрессия-ловушка: approve_match не проверяет, что matched_name
    заполнен — простое "Принять" на пустой позиции не должно делать её
    "проверено" в бейдже уверенности."""
    with app.app_context():
        order_id = _make_bare_repair_order(app)
        match = PartMatch(
            repair_order_id=order_id,
            contract_name="Деталь",
            matched_name=None,
            confidence_level=ConfidenceLevel.LLM_GUESS,
            confidence_score=0.0,
            review_status=ReviewStatus.PENDING,
            raw_match_data={"source": "no_match_found"},
        )
        db.session.add(match)
        db.session.commit()
        match_id = match.id

    resp = client.post(f"/api/repair-orders/matching/{match_id}/approve", headers=admin_headers)
    assert resp.get_json()["review_status"] == "approved"
    assert resp.get_json()["is_verified"] is False


def test_list_candidates_returns_contract_catalog_not_repair_order_lines(
    client, admin_headers, repair_order_with_matches, app
):
    """Регрессия: раньше ручной переподбор искал среди parsed_lines самого
    заказ-наряда (черновика мехника) — то есть среди того, что как раз
    нужно сопоставить, а не среди реального прайса договора с проверенными
    ценами. Источник кандидатов должен быть каталог договора (ContractPart)."""
    order_id = repair_order_with_matches["order_id"]
    with app.app_context():
        order = db.session.get(RepairOrder, order_id)
        db.session.add(
            ContractPart(contract_id=order.contract_id, article="ABC-1", name="Диск тормозной", price=1200.0)
        )
        db.session.commit()

    resp = client.get(f"/api/repair-orders/matching/{order_id}/candidates", headers=admin_headers)
    assert resp.status_code == 200
    assert resp.get_json() == [{"article": "ABC-1", "name": "Диск тормозной", "price": 1200.0}]


def test_list_candidates_filters_by_query_across_name_and_article(
    client, admin_headers, repair_order_with_matches, app
):
    order_id = repair_order_with_matches["order_id"]
    with app.app_context():
        order = db.session.get(RepairOrder, order_id)
        db.session.add_all(
            [
                ContractPart(contract_id=order.contract_id, article="ABC-1", name="Диск тормозной", price=1200.0),
                ContractPart(contract_id=order.contract_id, article="XYZ-9", name="Фильтр масляный", price=300.0),
            ]
        )
        db.session.commit()

    by_name = client.get(
        f"/api/repair-orders/matching/{order_id}/candidates?q=тормоз", headers=admin_headers
    ).get_json()
    assert [c["article"] for c in by_name] == ["ABC-1"]

    by_article = client.get(
        f"/api/repair-orders/matching/{order_id}/candidates?q=xyz", headers=admin_headers
    ).get_json()
    assert [c["article"] for c in by_article] == ["XYZ-9"]


def test_list_candidates_only_returns_parts_from_this_repair_orders_own_contract(
    client, admin_headers, repair_order_with_matches, app
):
    order_id = repair_order_with_matches["order_id"]
    with app.app_context():
        order = db.session.get(RepairOrder, order_id)
        other_contract = Contract(
            original_filename="other.xlsx", storage_path="/tmp/other.xlsx", status=DocumentProcessingStatus.PARSED
        )
        db.session.add(other_contract)
        db.session.flush()
        db.session.add(ContractPart(contract_id=other_contract.id, article="FOREIGN-1", name="Чужая деталь", price=1.0))
        db.session.add(ContractPart(contract_id=order.contract_id, article="OWN-1", name="Своя деталь", price=1.0))
        db.session.commit()

    resp = client.get(f"/api/repair-orders/matching/{order_id}/candidates", headers=admin_headers)
    articles = {c["article"] for c in resp.get_json()}
    assert articles == {"OWN-1"}


def test_edit_match_marks_manually_edited_and_approved(client, admin_headers, repair_order_with_matches, app):
    match_id = repair_order_with_matches["match_ids"][1]
    resp = client.patch(
        f"/api/repair-orders/matching/{match_id}",
        headers=admin_headers,
        json={"matched_article": "XYZ-9-ALT", "matched_name": "Фильтр альтернативный"},
    )
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["manually_edited"] is True
    assert body["review_status"] == "approved"
    assert body["matched_article"] == "XYZ-9-ALT"


def test_edit_match_requires_article_or_name(client, admin_headers, repair_order_with_matches):
    match_id = repair_order_with_matches["match_ids"][0]
    resp = client.patch(
        f"/api/repair-orders/matching/{match_id}", headers=admin_headers, json={}
    )
    assert resp.status_code == 400


def test_approve_and_reject_match(client, admin_headers, repair_order_with_matches):
    m1, m2 = repair_order_with_matches["match_ids"]

    approve_resp = client.post(
        f"/api/repair-orders/matching/{m1}/approve", headers=admin_headers
    )
    assert approve_resp.get_json()["review_status"] == "approved"

    reject_resp = client.post(f"/api/repair-orders/matching/{m2}/reject", headers=admin_headers)
    assert reject_resp.get_json()["review_status"] == "rejected"


def test_bulk_review(client, admin_headers, repair_order_with_matches):
    ids = repair_order_with_matches["match_ids"]
    resp = client.post(
        "/api/repair-orders/matching/bulk",
        headers=admin_headers,
        json={"ids": ids, "action": "approve"},
    )
    assert resp.status_code == 200
    assert all(m["review_status"] == "approved" for m in resp.get_json())


def test_bulk_review_validates_action(client, admin_headers, repair_order_with_matches):
    resp = client.post(
        "/api/repair-orders/matching/bulk",
        headers=admin_headers,
        json={"ids": repair_order_with_matches["match_ids"], "action": "explode"},
    )
    assert resp.status_code == 400


def test_export_csv(client, admin_headers, repair_order_with_matches):
    order_id = repair_order_with_matches["order_id"]
    resp = client.get(
        f"/api/repair-orders/matching/{order_id}/export", headers=admin_headers
    )
    assert resp.status_code == 200
    assert resp.mimetype == "text/csv"
    text = resp.get_data(as_text=True)
    assert "ABC-1" in text
    assert "XYZ-9" in text


def test_generate_document_blocked_while_pending(client, admin_headers, repair_order_with_matches):
    order_id = repair_order_with_matches["order_id"]
    resp = client.post(
        f"/api/repair-orders/matching/{order_id}/generate-document", headers=admin_headers
    )
    assert resp.status_code == 409


def test_generate_document_succeeds_once_all_reviewed(
    client, admin_headers, repair_order_with_matches, tmp_path, app
):
    order_id = repair_order_with_matches["order_id"]

    with app.app_context():
        order = db.session.get(RepairOrder, order_id)
        order.storage_path = str(tmp_path / "order.xlsx")
        db.session.commit()

    client.post(
        "/api/repair-orders/matching/bulk",
        headers=admin_headers,
        json={"ids": repair_order_with_matches["match_ids"], "action": "approve"},
    )

    resp = client.post(
        f"/api/repair-orders/matching/{order_id}/generate-document", headers=admin_headers
    )
    assert resp.status_code == 200
    assert resp.mimetype in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
    )

    with app.app_context():
        order = db.session.get(RepairOrder, order_id)
        # Регрессия: RepairOrderStatus.REVIEWED никогда не выставлялся —
        # заказ-наряд навсегда оставался needs_review даже после генерации
        # документа, засоряя счётчик "нужно проверить" на дашборде.
        assert order.status == RepairOrderStatus.REVIEWED


def test_generate_document_with_template_uses_uploaded_template(
    client, admin_headers, repair_order_with_matches, tmp_path, app
):
    import io

    import openpyxl

    order_id = repair_order_with_matches["order_id"]

    with app.app_context():
        order = db.session.get(RepairOrder, order_id)
        order.storage_path = str(tmp_path / "order.xlsx")
        db.session.commit()

    client.post(
        "/api/repair-orders/matching/bulk",
        headers=admin_headers,
        json={"ids": repair_order_with_matches["match_ids"], "action": "approve"},
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Заказ-наряд № {{order_number}}"])
    ws.append(["{{part.n}}", "{{part.article}}", "{{part.price}}"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    upload_resp = client.post(
        "/api/document-templates",
        headers=admin_headers,
        data={"name": "Кастомный", "file": (buf, "custom.xlsx")},
        content_type="multipart/form-data",
    )
    template_id = upload_resp.get_json()["id"]

    resp = client.post(
        f"/api/repair-orders/matching/{order_id}/generate-document?template_id={template_id}",
        headers=admin_headers,
    )
    assert resp.status_code == 200

    output_wb = openpyxl.load_workbook(io.BytesIO(resp.data))
    values = [cell.value for row in output_wb.active.iter_rows() for cell in row if cell.value is not None]
    assert f"Заказ-наряд № {order_id}" in values
    assert "ABC-1" in values


def test_add_part_creates_new_approved_match(client, admin_headers, repair_order_with_matches):
    order_id = repair_order_with_matches["order_id"]
    resp = client.post(
        f"/api/repair-orders/matching/{order_id}/parts",
        headers=admin_headers,
        json={"matched_article": "290 074", "matched_name": "Амортизатор Sachs", "matched_price": 10920.33, "source": "rossco"},
    )
    assert resp.status_code == 201
    body = resp.get_json()
    assert body["matched_article"] == "290 074"
    assert body["matched_price"] == 10920.33
    assert body["review_status"] == "approved"
    assert body["confidence_level"] == "exact"
    assert body["manually_edited"] is True

    listed = client.get(f"/api/repair-orders/matching/{order_id}", headers=admin_headers)
    assert len(listed.get_json()) == 3


def test_add_part_requires_matched_name(client, admin_headers, repair_order_with_matches):
    order_id = repair_order_with_matches["order_id"]
    resp = client.post(
        f"/api/repair-orders/matching/{order_id}/parts",
        headers=admin_headers,
        json={"matched_article": "290 074"},
    )
    assert resp.status_code == 400


def test_add_part_unknown_repair_order_404(client, admin_headers):
    resp = client.post(
        "/api/repair-orders/matching/99999/parts", headers=admin_headers, json={"matched_name": "x"}
    )
    assert resp.status_code == 404
