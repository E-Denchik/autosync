import openpyxl

from app.extensions import db
from app.models import (
    Contract,
    ConfidenceLevel,
    Contragent,
    DocumentProcessingStatus,
    LaborLine,
    PartMatch,
    RepairOrder,
    RepairOrderStatus,
    ReviewStatus,
)
from app.services import company_profile
from app.services.document_generator import build_template_context, generate_repair_order_document


def _make_repair_order(app, tmp_path, **overrides) -> RepairOrder:
    contract = Contract(original_filename="c.xlsx", storage_path="/tmp/c.xlsx", status=DocumentProcessingStatus.PARSED)
    db.session.add(contract)
    db.session.flush()
    defaults = dict(
        contract_id=contract.id,
        original_filename="o.xlsx",
        storage_path=str(tmp_path / "o.xlsx"),
        status=RepairOrderStatus.NEEDS_REVIEW,
        vehicle_make="KIA",
        vehicle_model="Rio",
        vehicle_vin="TESTVIN12345",
        vehicle_year=2020,
    )
    defaults.update(overrides)
    order = RepairOrder(**defaults)
    db.session.add(order)
    db.session.commit()
    return order


def _rows(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    return [tuple(row) for row in ws.iter_rows(values_only=True)]


def test_generates_readable_xlsx_with_correct_totals(app, tmp_path):
    with app.app_context():
        order = _make_repair_order(app, tmp_path)

        db.session.add(
            PartMatch(
                repair_order_id=order.id,
                contract_article="A-1",
                contract_name="Деталь черновика",
                matched_article="A-1",
                matched_name="Тормозной диск",
                matched_price=1500,
                confidence_level=ConfidenceLevel.EXACT,
                review_status=ReviewStatus.APPROVED,
                nomenclature_cat_number="CAT-1",
                nomenclature_manufacturer="LUZAR",
                nomenclature_unit="шт",
                nomenclature_warehouse="Основной",
            )
        )
        db.session.add(
            LaborLine(
                repair_order_id=order.id,
                description="Замена колодок",
                matched_operation_name="Замена тормозных колодок",
                norm_hours=1.5,
                hourly_rate=1000,
                total_cost=1500,
                confidence_level=ConfidenceLevel.LLM_GUESS,
                review_status=ReviewStatus.APPROVED,
            )
        )
        db.session.commit()

        path = generate_repair_order_document(order)

    rows = _rows(path)
    flat = [cell for row in rows for cell in row if cell is not None]

    assert any(f"Заказ-наряд № {order.id}" in str(c) for c in flat)
    assert any("KIA Rio" in str(c) for c in flat)
    assert any("TESTVIN12345" in str(c) for c in flat)

    # Работы: строка с деталями операции, норма-часами и суммой.
    # openpyxl дополняет каждую строку до ширины листа значениями None при
    # чтении, поэтому сверяем начало строки срезом, а не строку целиком.
    labor_row = next(r for r in rows if r[1] == "Замена тормозных колодок")
    assert labor_row[:5] == (1, "Замена тормозных колодок", 1.5, 1000.0, 1500.0)
    assert next(r for r in rows if r[3] == "Итого работы:")[:5] == (None, None, None, "Итого работы:", 1500.0)

    # Запчасти: артикул/номенклатура/кол-во/цена/сумма/склад из PartMatch.
    part_row = next(r for r in rows if r[1] == "A-1")
    assert part_row[:10] == (1, "A-1", "CAT-1", "Тормозной диск", "LUZAR", "шт", 1.0, 1500.0, 1500.0, "Основной")
    assert next(r for r in rows if r[7] == "Итого запчасти:")[:9] == (
        None, None, None, None, None, None, None, "Итого запчасти:", 1500.0,
    )

    assert next(r for r in rows if r[7] == "ИТОГО:")[:9] == (
        None, None, None, None, None, None, None, "ИТОГО:", 3000.0,
    )


def test_quantity_multiplies_price_in_totals(app, tmp_path):
    """Регрессия: 2 шт. по 1500 ₽ должны дать 3000 ₽ в итоговом документе,
    а не 1500 ₽ — раньше contract_qty вообще не сохранялся и не участвовал
    в расчёте суммы (см. PartMatch.contract_qty)."""
    with app.app_context():
        order = _make_repair_order(app, tmp_path)
        db.session.add(
            PartMatch(
                repair_order_id=order.id,
                contract_article="A-1",
                contract_name="Тормозные колодки",
                contract_qty=2,
                matched_article="A-1",
                matched_name="Тормозные колодки",
                matched_price=1500,
                confidence_level=ConfidenceLevel.EXACT,
                review_status=ReviewStatus.APPROVED,
            )
        )
        db.session.commit()

        path = generate_repair_order_document(order)
        context, part_items, _ = build_template_context(order)

    part_row = next(r for r in _rows(path) if r[1] == "A-1")
    assert part_row[6] == 2.0  # Кол-во
    assert part_row[7] == 1500.0  # Цена (за единицу)
    assert part_row[8] == 3000.0  # Сумма

    assert any(r[:9] == (None, None, None, None, None, None, None, "Итого запчасти:", 3000.0) for r in _rows(path))

    assert context["parts_total"] == 3000.0
    assert part_items[0]["qty"] == 2.0
    assert part_items[0]["total"] == 3000.0


def test_generate_document_neutralizes_formula_injection_in_part_name(app, tmp_path):
    """Регрессия: matched_name/matched_article приходят из данных
    заказ-наряда/договора (в конечном счёте — из загруженного заказчиком
    файла), не из доверенной формы — строка с ведущим "=" попадала в ячейку
    как есть, а openpyxl сам помечает такую ячейку как формулу, которая
    выполнится при открытии готового документа. См. app/services/xlsx_safety.py."""
    with app.app_context():
        order = _make_repair_order(app, tmp_path)
        db.session.add(
            PartMatch(
                repair_order_id=order.id,
                contract_article="A-1",
                contract_name="Деталь",
                matched_article="=1+1",
                matched_name="=CMD('/c calc')A1",
                matched_price=100,
                confidence_level=ConfidenceLevel.EXACT,
                review_status=ReviewStatus.APPROVED,
            )
        )
        db.session.commit()

        path = generate_repair_order_document(order)

    wb = openpyxl.load_workbook(path)
    article_cell = next(c for row in wb.active.iter_rows() for c in row if c.value == "'=1+1")
    name_cell = next(c for row in wb.active.iter_rows() for c in row if c.value == "'=CMD('/c calc')A1")
    assert article_cell.data_type != "f"
    assert name_cell.data_type != "f"


def test_excludes_pending_and_rejected_matches(app, tmp_path):
    with app.app_context():
        order = _make_repair_order(app, tmp_path)

        db.session.add(
            PartMatch(
                repair_order_id=order.id,
                contract_name="Одобренная деталь",
                matched_name="Одобренная деталь",
                matched_price=100,
                confidence_level=ConfidenceLevel.EXACT,
                review_status=ReviewStatus.APPROVED,
            )
        )
        db.session.add(
            PartMatch(
                repair_order_id=order.id,
                contract_name="Ожидающая деталь",
                matched_name="Ожидающая деталь",
                matched_price=200,
                confidence_level=ConfidenceLevel.LLM_GUESS,
                review_status=ReviewStatus.PENDING,
            )
        )
        db.session.add(
            PartMatch(
                repair_order_id=order.id,
                contract_name="Отклонённая деталь",
                matched_name="Отклонённая деталь",
                matched_price=300,
                confidence_level=ConfidenceLevel.LLM_GUESS,
                review_status=ReviewStatus.REJECTED,
            )
        )
        db.session.commit()

        path = generate_repair_order_document(order)

    flat = [cell for row in _rows(path) for cell in row if cell is not None]
    assert any("Одобренная деталь" in str(c) for c in flat)
    assert not any("Ожидающая деталь" in str(c) for c in flat)
    assert not any("Отклонённая деталь" in str(c) for c in flat)
    assert any(
        r[:9] == (None, None, None, None, None, None, None, "Итого запчасти:", 100.0) for r in _rows(path)
    )


def test_includes_company_profile_when_set(app, tmp_path):
    with app.app_context():
        company_profile.save(
            {
                "COMPANY_NAME": "ООО Тестовый Автосервис",
                "COMPANY_INN": "1234567890",
                "COMPANY_ADDRESS": "г. Тест, ул. Примерная, 1",
                "COMPANY_PHONE": "+7 900 000-00-00",
            }
        )
        order = _make_repair_order(app, tmp_path)
        path = generate_repair_order_document(order)

    flat = [cell for row in _rows(path) for cell in row if cell is not None]
    assert any("ООО Тестовый Автосервис" in str(c) for c in flat)
    assert any("1234567890" in str(c) for c in flat)


def test_omits_company_profile_block_when_not_set(app, tmp_path):
    with app.app_context():
        order = _make_repair_order(app, tmp_path)
        path = generate_repair_order_document(order)

    flat = [cell for row in _rows(path) for cell in row if cell is not None]
    # Ничего похожего на реквизиты компании не появляется, если профиль не задан.
    assert not any("ИНН" in str(c) for c in flat)
    assert any(str(c).startswith("Заказ-наряд №") for c in flat)


def test_includes_contragent_name_when_present(app, tmp_path):
    with app.app_context():
        contragent = Contragent(name="СТО Восток", hourly_rate=1000)
        db.session.add(contragent)
        db.session.flush()
        order = _make_repair_order(app, tmp_path, contragent_id=contragent.id)
        path = generate_repair_order_document(order)

    flat = [cell for row in _rows(path) for cell in row if cell is not None]
    assert any("СТО Восток" in str(c) for c in flat)


def test_zero_matches_gives_zero_totals_not_crash(app, tmp_path):
    with app.app_context():
        order = _make_repair_order(app, tmp_path)
        path = generate_repair_order_document(order)

    rows = _rows(path)
    assert any(r[:9] == (None, None, None, None, None, None, None, "ИТОГО:", 0.0) for r in rows)


def test_build_template_context_computes_totals_and_items(app, tmp_path):
    with app.app_context():
        order = _make_repair_order(app, tmp_path)
        db.session.add(
            PartMatch(
                repair_order_id=order.id,
                contract_article="A-1",
                matched_article="A-1",
                matched_name="Деталь",
                matched_price=500,
                confidence_level=ConfidenceLevel.EXACT,
                review_status=ReviewStatus.APPROVED,
                nomenclature_manufacturer="LUZAR",
            )
        )
        db.session.add(
            LaborLine(
                repair_order_id=order.id,
                description="Работа",
                matched_operation_name="Работа",
                norm_hours=2,
                hourly_rate=800,
                total_cost=1600,
                confidence_level=ConfidenceLevel.LLM_GUESS,
                review_status=ReviewStatus.APPROVED,
            )
        )
        db.session.commit()

        context, part_items, labor_items = build_template_context(order)

    assert context["parts_total"] == 500.0
    assert context["labor_total"] == 1600.0
    assert context["grand_total"] == 2100.0
    assert context["vehicle_make"] == "KIA"
    assert context["order_number"] == order.id

    assert part_items == [
        {
            "article": "A-1",
            "cat_number": "",
            "name": "Деталь",
            "manufacturer": "LUZAR",
            "unit": "",
            "qty": 1.0,
            "price": 500.0,
            "total": 500.0,
            "warehouse": "",
        }
    ]
    assert labor_items == [
        {"description": "Работа", "norm_hours": 2.0, "hourly_rate": 800.0, "total": 1600.0}
    ]


def test_uses_real_order_number_and_date_from_source_file_when_available(app, tmp_path):
    """Регрессия: order_number/order_date распознаются из самого файла при
    загрузке (см. document_parser.parse_repair_order_export), но раньше
    никуда не сохранялись — итоговый документ вместо реального номера/даты
    заказ-наряда подставлял внутренний id записи и дату ЗАГРУЗКИ в систему,
    которые не совпадают с тем, что было написано в файле у заказчика."""
    with app.app_context():
        order = _make_repair_order(app, tmp_path, order_number="0000010749", order_date="14.01.2026")
        path = generate_repair_order_document(order)

    flat = [cell for row in _rows(path) for cell in row if cell is not None]
    assert any("Заказ-наряд № 0000010749 от 14.01.2026" in str(c) for c in flat)
    assert not any(f"№ {order.id} от" in str(c) for c in flat)

    context, _, _ = build_template_context(order)
    assert context["order_number"] == "0000010749"
    assert context["order_date"] == "14.01.2026"


def test_falls_back_to_id_and_created_at_when_order_number_not_recognized(app, tmp_path):
    """Не у каждого файла получается распознать номер/дату (не 1С-формат,
    сканы и т.п.) — тогда, как и раньше, используются id записи и дата
    загрузки, а не пустая строка в документе."""
    with app.app_context():
        order = _make_repair_order(app, tmp_path)  # order_number/order_date не заданы
        path = generate_repair_order_document(order)

    flat = [cell for row in _rows(path) for cell in row if cell is not None]
    assert any(f"Заказ-наряд № {order.id} от {order.created_at.strftime('%d.%m.%Y')}" in str(c) for c in flat)


def test_build_template_context_approved_only_false_includes_pending(app, tmp_path):
    with app.app_context():
        order = _make_repair_order(app, tmp_path)
        db.session.add(
            PartMatch(
                repair_order_id=order.id,
                contract_name="Ожидает",
                matched_name="Ожидает",
                matched_price=50,
                confidence_level=ConfidenceLevel.LLM_GUESS,
                review_status=ReviewStatus.PENDING,
            )
        )
        db.session.commit()

        context, part_items, _ = build_template_context(order, approved_only=False)

    assert context["parts_total"] == 50.0
    assert len(part_items) == 1
