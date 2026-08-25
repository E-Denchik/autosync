import openpyxl

from app.services.document_template_engine import build_starter_template, render_template


def _build_template(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Заказ-наряд № {{order_number}}"])
    ws.append(["Заказчик: {{client_name}}"])
    ws.append(["№", "Работа", "Сумма"])
    ws.append(["{{labor.n}}", "{{labor.description}}", "{{labor.total}}"])
    ws.append(["№", "Артикул", "Цена"])
    ws.append(["{{part.n}}", "{{part.article}}", "{{part.price}}"])
    ws.append(["ИТОГО:", "", "{{grand_total}}"])
    wb.save(path)


def test_render_template_substitutes_header_tokens(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _build_template(template_path)

    render_template(
        str(template_path),
        str(output_path),
        {"order_number": 42, "client_name": "ООО Ромашка", "grand_total": 1500},
        part_items=[],
        labor_items=[],
    )

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    assert ws["A1"].value == "Заказ-наряд № 42"
    assert ws["A2"].value == "Заказчик: ООО Ромашка"


def test_render_template_expands_repeating_rows_for_multiple_items(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _build_template(template_path)

    render_template(
        str(template_path),
        str(output_path),
        {"order_number": 1, "client_name": "x", "grand_total": 100},
        part_items=[
            {"article": "A-1", "price": 100},
            {"article": "A-2", "price": 200},
            {"article": "A-3", "price": 300},
        ],
        labor_items=[{"description": "Замена масла", "total": 500}],
    )

    wb = openpyxl.load_workbook(output_path)
    rows = list(ws_rows(wb.active))
    parts_rows = [r for r in rows if r[1] in ("A-1", "A-2", "A-3")]
    assert len(parts_rows) == 3
    assert any(r[1] == "A-1" and r[2] == 100 for r in parts_rows)
    assert any(r[1] == "A-2" and r[2] == 200 for r in parts_rows)
    assert any(r[1] == "A-3" and r[2] == 300 for r in parts_rows)
    assert any(r[1] == "Замена масла" and r[2] == 500 for r in rows)


def test_render_template_deletes_row_when_no_items(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _build_template(template_path)

    render_template(
        str(template_path),
        str(output_path),
        {"order_number": 1, "client_name": "x", "grand_total": 0},
        part_items=[],
        labor_items=[],
    )

    wb = openpyxl.load_workbook(output_path)
    values = [cell.value for row in wb.active.iter_rows() for cell in row if cell.value is not None]
    assert not any(isinstance(v, str) and "{{part" in v for v in values)
    assert not any(isinstance(v, str) and "{{labor" in v for v in values)


def ws_rows(ws):
    for row in ws.iter_rows(values_only=True):
        yield list(row)


def test_render_template_reports_malformed_tokens_left_unresolved(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Заказ-наряд № {{order_number}}"])
    ws.append(["Дата: {{order date}}"])
    wb.save(template_path)

    _, unresolved = render_template(
        str(template_path),
        str(output_path),
        {"order_number": 42},
        part_items=[],
        labor_items=[],
    )

    assert unresolved == ["{{order date}}"]
    wb_out = openpyxl.load_workbook(output_path)
    assert wb_out.active["A1"].value == "Заказ-наряд № 42"
    assert wb_out.active["A2"].value == "Дата: {{order date}}"


def test_merged_cells_below_repeating_rows_do_not_swallow_data(tmp_path):
    """Регрессия: реальные шаблоны заказчика почти всегда содержат
    объединённые ячейки (заголовок, строка "ИТОГО" и т.п.). openpyxl
    insert_rows/delete_rows НЕ сдвигает merged_cells — строка "ИТОГО",
    объединённая ниже {{part.*}}, оставалась на старых координатах после
    вставки строк под несколько позиций, и запись данных во "вторую"
    (не главную) ячейку сдвинувшегося объединения openpyxl тихо
    игнорирует — часть заказ-наряда (например, артикул) пропадала бы из
    документа без единой ошибки, только сдвинутый на старое место "хвост"
    объединения."""
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Заказ-наряд № {{order_number}}"])
    ws.merge_cells("A1:C1")
    ws.append(["№", "Артикул", "Цена"])
    ws.append(["{{part.n}}", "{{part.article}}", "{{part.price}}"])
    ws.append(["ИТОГО:", "", "{{grand_total}}"])
    ws.merge_cells("A4:B4")
    wb.save(template_path)

    render_template(
        str(template_path),
        str(output_path),
        {"order_number": 1, "grand_total": 600},
        part_items=[
            {"article": "A-1", "price": 100},
            {"article": "A-2", "price": 200},
            {"article": "A-3", "price": 300},
        ],
        labor_items=[],
    )

    wb_out = openpyxl.load_workbook(output_path)
    rows = list(ws_rows(wb_out.active))
    parts_rows = {r[1]: r[2] for r in rows if r[1] and r[1].startswith("A-")}
    assert parts_rows == {"A-1": 100, "A-2": 200, "A-3": 300}  # ни один артикул не потерялся

    itogo_row = next(r for r in rows if r[0] == "ИТОГО:")
    assert itogo_row[2] == 600

    # Заголовок (не задет вставкой строк) остался объединён как был; "ИТОГО"
    # сдвинулся вниз на 2 вставленные строки и объединён на новом месте.
    ranges = {str(r) for r in wb_out.active.merged_cells.ranges}
    assert "A1:C1" in ranges
    assert "A6:B6" in ranges
    assert "A4:B4" not in ranges


def test_merged_cells_below_deleted_row_shift_up_when_no_items(tmp_path):
    """Тот же класс бага, но для случая delete_rows (0 позиций) — merge ниже
    удаляемой строки тоже должен сдвинуться, а не остаться на старом месте."""
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Заказ-наряд № {{order_number}}"])
    ws.append(["№", "Артикул", "Цена"])
    ws.append(["{{part.n}}", "{{part.article}}", "{{part.price}}"])
    ws.append(["ИТОГО:", "", "{{grand_total}}"])
    ws.merge_cells("A4:B4")
    wb.save(template_path)

    render_template(
        str(template_path),
        str(output_path),
        {"order_number": 1, "grand_total": 0},
        part_items=[],
        labor_items=[],
    )

    wb_out = openpyxl.load_workbook(output_path)
    rows = list(ws_rows(wb_out.active))
    itogo_row = next(r for r in rows if r[0] == "ИТОГО:")
    assert itogo_row[2] == 0

    ranges = {str(r) for r in wb_out.active.merged_cells.ranges}
    assert "A3:B3" in ranges  # сдвинулось на 1 строку вверх после удаления шаблонной строки
    assert "A4:B4" not in ranges


def test_build_starter_template_has_no_leftover_placeholders_after_render(tmp_path):
    template_path = tmp_path / "starter.xlsx"
    output_path = tmp_path / "output.xlsx"
    build_starter_template(str(template_path))

    render_template(
        str(template_path),
        str(output_path),
        {
            "company_name": "ИП Иванов",
            "company_inn": "123",
            "company_address": "г. Самара",
            "company_phone": "+7",
            "order_number": 1,
            "order_date": "01.01.2026",
            "client_name": "Клиент",
            "vehicle_make": "LADA",
            "vehicle_model": "Granta",
            "vehicle_vin": "VIN123",
            "vehicle_year": 2020,
            "parts_total": 100,
            "labor_total": 200,
            "grand_total": 300,
        },
        part_items=[{"n": 1, "article": "A-1", "cat_number": "", "name": "Деталь", "manufacturer": "", "unit": "", "price": 100, "warehouse": ""}],
        labor_items=[{"n": 1, "description": "Работа", "norm_hours": 1, "hourly_rate": 200, "total": 200}],
    )

    wb = openpyxl.load_workbook(output_path)
    values = [cell.value for row in wb.active.iter_rows() for cell in row if cell.value is not None]
    assert not any(isinstance(v, str) and "{{" in v for v in values)
