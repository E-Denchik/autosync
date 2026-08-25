from __future__ import annotations

import re
from copy import copy

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.services.xlsx_safety import sanitize_cell_value

TOKEN_RE = re.compile(r"\{\{\s*([\w.]+)\s*\}\}")
FULL_TOKEN_RE = re.compile(r"^\{\{\s*([\w.]+)\s*\}\}$")
LEFTOVER_TOKEN_RE = re.compile(r"\{\{[^{}]*\}\}")


class DocumentTemplateError(RuntimeError):
    pass


def _substitute_cell(cell, values: dict) -> None:
    text = cell.value
    full_match = FULL_TOKEN_RE.match(text)
    if full_match:
        cell.value = sanitize_cell_value(values.get(full_match.group(1)))
        return

    def repl(m: re.Match) -> str:
        value = values.get(m.group(1))
        return "" if value is None else str(value)

    cell.value = sanitize_cell_value(TOKEN_RE.sub(repl, text))


def _find_row(ws, prefix: str) -> int | None:
    marker = "{{" + prefix + "."
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and marker in cell.value:
                return cell.row
    return None


def _copy_row(ws, src_row: int, dst_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        dst.value = src.value
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)


def _expand_rows(ws, prefix: str, items: list[dict]) -> tuple[int, int] | None:
    """Возвращает (номер_строки_шаблона, дельта) — на сколько строк сдвинулось
    всё, что было ниже неё (+N вставлено, -1 удалена) — или None, если
    строка с "{{prefix." в шаблоне не найдена вовсе. Вызывающий код
    использует это, чтобы потом восстановить объединённые ячейки на
    правильных координатах (см. render_template)."""
    template_row = _find_row(ws, prefix)
    if template_row is None:
        return None
    if not items:
        ws.delete_rows(template_row, 1)
        return (template_row, -1)

    extra = len(items) - 1
    if extra > 0:
        ws.insert_rows(template_row + 1, extra)
        for i in range(extra):
            _copy_row(ws, template_row, template_row + 1 + i)

    for i, item in enumerate(items):
        row_idx = template_row + i
        values = {f"{prefix}.{k}": v for k, v in item.items()}
        values[f"{prefix}.n"] = i + 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col)
            if isinstance(cell.value, str):
                _substitute_cell(cell, values)

    return (template_row, extra)


def _reapply_shifted_merges(ws, original_ranges: list[tuple[int, int, int, int]], shifts: list[tuple[int, int]]) -> None:
    """original_ranges — границы объединений ДО вставки/удаления строк
    (min_row, min_col, max_row, max_col). shifts — в порядке применения:
    (строка_шаблона, дельта). Диапазон, задевающий саму строку-шаблон
    ({{part.*}}/{{labor.*}}), не восстанавливаем — после раскрытия в
    несколько независимых строк данных единого прямоугольника для него уже
    нет; всё, что было строго ниже, просто сдвигаем на дельту."""
    for min_row, min_col, max_row, max_col in original_ranges:
        cur_min, cur_max = min_row, max_row
        dropped = False
        for affected_row, delta in shifts:
            if cur_min <= affected_row <= cur_max:
                dropped = True
                break
            if cur_min > affected_row:
                cur_min += delta
                cur_max += delta
        if dropped or cur_max < cur_min or cur_min < 1:
            continue
        rng = f"{get_column_letter(min_col)}{cur_min}:{get_column_letter(max_col)}{cur_max}"
        try:
            ws.merge_cells(rng)
        except ValueError:
            pass


def render_template(
    template_path: str,
    output_path: str,
    context: dict,
    part_items: list[dict],
    labor_items: list[dict],
) -> tuple[str, list[str]]:
    try:
        wb = openpyxl.load_workbook(template_path)
    except Exception as exc:
        raise DocumentTemplateError(f"Не удалось открыть файл шаблона: {exc}") from exc

    ws = wb.active

    # openpyxl insert_rows/delete_rows не двигает объединённые ячейки — без
    # этого шага строка "ИТОГО" (или любая другая) с merge ниже
    # {{part.*}}/{{labor.*}} осталась бы на СТАРЫХ координатах после вставки
    # строк под несколько позиций, а запись данных в "не главную" ячейку
    # чужого (сдвинувшегося относительно неё) объединения openpyxl тихо
    # игнорирует — часть строки заказ-наряда молча пропадала бы из
    # документа без единой ошибки. Снимаем все объединения перед вставкой,
    # восстанавливаем на пересчитанных координатах в конце.
    original_merges = [(r.min_row, r.min_col, r.max_row, r.max_col) for r in list(ws.merged_cells.ranges)]
    for r in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(r))

    shifts: list[tuple[int, int]] = []
    part_shift = _expand_rows(ws, "part", part_items)
    if part_shift is not None:
        shifts.append(part_shift)
    labor_shift = _expand_rows(ws, "labor", labor_items)
    if labor_shift is not None:
        shifts.append(labor_shift)

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "{{" in cell.value:
                _substitute_cell(cell, context)

    _reapply_shifted_merges(ws, original_merges, shifts)

    unresolved = set()
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                unresolved.update(LEFTOVER_TOKEN_RE.findall(cell.value))

    wb.save(output_path)
    return output_path, sorted(unresolved)


def build_starter_template(output_path: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказ-наряд"
    bold = Font(bold=True)

    def row(*values):
        ws.append(list(values))

    row("{{company_name}}")
    ws["A1"].font = bold
    row("ИНН {{company_inn}}   Адрес: {{company_address}}   Тел: {{company_phone}}")
    row()
    row("Заказ-наряд № {{order_number}} от {{order_date}}")
    row()
    row("Заказчик: {{client_name}}")
    row("Автомобиль: {{vehicle_make}} {{vehicle_model}}   VIN: {{vehicle_vin}}   {{vehicle_year}} г.")
    row()

    row("Выполненные работы")
    ws[f"A{ws.max_row}"].font = bold
    row("№", "Работа", "Норма, ч", "Цена н/ч", "Сумма")
    row("{{labor.n}}", "{{labor.description}}", "{{labor.norm_hours}}", "{{labor.hourly_rate}}", "{{labor.total}}")
    row()
    row("", "", "", "Итого работы:", "{{labor_total}}")
    ws[f"D{ws.max_row}"].font = bold
    row()

    row("Запчасти и материалы")
    ws[f"A{ws.max_row}"].font = bold
    row("№", "Артикул", "№ кат.", "Наименование", "Производитель", "Ед.", "Кол-во", "Цена", "Сумма", "Склад")
    row(
        "{{part.n}}",
        "{{part.article}}",
        "{{part.cat_number}}",
        "{{part.name}}",
        "{{part.manufacturer}}",
        "{{part.unit}}",
        "{{part.qty}}",
        "{{part.price}}",
        "{{part.total}}",
        "{{part.warehouse}}",
    )
    row()
    row("", "", "", "", "", "", "", "Итого запчасти:", "{{parts_total}}")
    ws[f"H{ws.max_row}"].font = bold
    row()

    row("", "", "", "", "", "", "", "ИТОГО:", "{{grand_total}}")
    ws[f"H{ws.max_row}"].font = bold
    ws[f"I{ws.max_row}"].font = bold

    for col_letter, width in {
        "A": 14, "B": 22, "C": 14, "D": 30, "E": 20, "F": 10, "G": 10, "H": 16, "I": 16, "J": 14,
    }.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(output_path)
    return output_path
