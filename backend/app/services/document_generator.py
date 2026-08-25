"""Генерация итогового заказ-наряда (xlsx) с подставленными ценами
из проверенных сопоставлений PartMatch и LaborLine."""

from __future__ import annotations

import os

import openpyxl
from openpyxl.styles import Font

from app.models import DocumentTemplate, LaborLine, PartMatch, RepairOrder, ReviewStatus
from app.services import company_profile, document_template_engine
from app.services.xlsx_safety import sanitize_cell_value as _s


def generate_repair_order_document(repair_order: RepairOrder) -> str:
    part_matches = (
        PartMatch.query.filter_by(repair_order_id=repair_order.id)
        .filter(PartMatch.review_status == ReviewStatus.APPROVED)
        .order_by(PartMatch.id)
        .all()
    )
    labor_lines = (
        LaborLine.query.filter_by(repair_order_id=repair_order.id)
        .filter(LaborLine.review_status == ReviewStatus.APPROVED)
        .order_by(LaborLine.id)
        .all()
    )
    profile = company_profile.load()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказ-наряд"
    bold = Font(bold=True)

    if profile["COMPANY_NAME"]:
        ws.append([profile["COMPANY_NAME"]])
        ws[f"A{ws.max_row}"].font = bold
    details = " ".join(
        part
        for part in [
            f"ИНН {profile['COMPANY_INN']}" if profile["COMPANY_INN"] else "",
            f"Адрес: {profile['COMPANY_ADDRESS']}" if profile["COMPANY_ADDRESS"] else "",
            f"Тел: {profile['COMPANY_PHONE']}" if profile["COMPANY_PHONE"] else "",
        ]
        if part
    )
    if details:
        ws.append([details])
    ws.append([])

    # Номер/дата исходного наряда, если их удалось распознать при загрузке
    # (см. RepairOrder.order_number/order_date) — иначе, как и раньше, id
    # записи и дата загрузки (не всегда совпадают с тем, что было в файле
    # у заказчика, но это лучше, чем ничего).
    order_number = repair_order.order_number or str(repair_order.id)
    order_date = repair_order.order_date or repair_order.created_at.strftime("%d.%m.%Y")
    ws.append([f"Заказ-наряд № {order_number} от {order_date}"])
    ws[f"A{ws.max_row}"].font = bold
    ws.append([])
    if repair_order.contragent:
        ws.append([f"Заказчик: {repair_order.contragent.name}"])
    vehicle = " ".join(filter(None, [repair_order.vehicle_make, repair_order.vehicle_model]))
    vehicle_line = f"Автомобиль: {vehicle}".strip()
    if repair_order.vehicle_vin:
        vehicle_line += f"   VIN: {repair_order.vehicle_vin}"
    if repair_order.vehicle_year:
        vehicle_line += f"   {repair_order.vehicle_year} г."
    ws.append([vehicle_line])
    ws.append([])

    ws.append(["Выполненные работы по заказ-наряду"])
    ws[f"A{ws.max_row}"].font = bold
    ws.append(["№", "Работа", "Норма, ч", "Цена н/ч", "Сумма"])

    labor_total = 0.0
    for i, line in enumerate(labor_lines, start=1):
        cost = float(line.total_cost) if line.total_cost is not None else 0.0
        labor_total += cost
        ws.append(
            [
                i,
                _s(line.matched_operation_name or line.description),
                float(line.norm_hours) if line.norm_hours is not None else "",
                float(line.hourly_rate) if line.hourly_rate is not None else "",
                cost or "",
            ]
        )
    ws.append(["", "", "", "Итого работы:", labor_total])
    ws[f"D{ws.max_row}"].font = bold
    ws.append([])

    ws.append(["Расходная накладная к заказ-наряду"])
    ws[f"A{ws.max_row}"].font = bold
    ws.append(["№", "Артикул", "№ кат.", "Наименование", "Производитель", "Ед.", "Кол-во", "Цена", "Сумма", "Склад"])

    parts_total = 0.0
    for i, match in enumerate(part_matches, start=1):
        price = float(match.matched_price) if match.matched_price is not None else 0.0
        qty = float(match.contract_qty) if match.contract_qty is not None else 1.0
        line_total = price * qty
        parts_total += line_total
        ws.append(
            [
                i,
                _s(match.matched_article or match.contract_article or ""),
                _s(match.nomenclature_cat_number or ""),
                _s(match.matched_name or match.contract_name or ""),
                _s(match.nomenclature_manufacturer or ""),
                _s(match.nomenclature_unit or ""),
                qty,
                price or "",
                line_total or "",
                _s(match.nomenclature_warehouse or ""),
            ]
        )
    ws.append(["", "", "", "", "", "", "", "Итого запчасти:", parts_total])
    ws[f"H{ws.max_row}"].font = bold
    ws.append([])

    ws.append(["", "", "", "", "", "", "", "ИТОГО:", parts_total + labor_total])
    ws[f"H{ws.max_row}"].font = bold
    ws[f"I{ws.max_row}"].font = bold

    for col_letter, width in {
        "A": 6, "B": 22, "C": 14, "D": 32, "E": 20, "F": 10, "G": 9, "H": 14, "I": 14, "J": 14,
    }.items():
        ws.column_dimensions[col_letter].width = width

    output_dir = os.path.dirname(repair_order.storage_path)
    output_path = os.path.join(output_dir, f"repair_order_{repair_order.id}_final.xlsx")
    wb.save(output_path)
    return output_path


def build_template_context(repair_order: RepairOrder, *, approved_only: bool = True) -> tuple[dict, list[dict], list[dict]]:
    part_query = PartMatch.query.filter_by(repair_order_id=repair_order.id)
    labor_query = LaborLine.query.filter_by(repair_order_id=repair_order.id)
    if approved_only:
        part_query = part_query.filter(PartMatch.review_status == ReviewStatus.APPROVED)
        labor_query = labor_query.filter(LaborLine.review_status == ReviewStatus.APPROVED)
    part_matches = part_query.order_by(PartMatch.id).all()
    labor_lines = labor_query.order_by(LaborLine.id).all()
    profile = company_profile.load()

    parts_total = sum(
        (float(m.matched_price) if m.matched_price is not None else 0.0)
        * (float(m.contract_qty) if m.contract_qty is not None else 1.0)
        for m in part_matches
    )
    labor_total = sum(float(l.total_cost) if l.total_cost is not None else 0.0 for l in labor_lines)

    context = {
        "company_name": profile["COMPANY_NAME"],
        "company_inn": profile["COMPANY_INN"],
        "company_address": profile["COMPANY_ADDRESS"],
        "company_phone": profile["COMPANY_PHONE"],
        "order_number": repair_order.order_number or repair_order.id,
        "order_date": repair_order.order_date or repair_order.created_at.strftime("%d.%m.%Y"),
        "client_name": repair_order.contragent.name if repair_order.contragent else "",
        "vehicle_make": repair_order.vehicle_make or "",
        "vehicle_model": repair_order.vehicle_model or "",
        "vehicle_vin": repair_order.vehicle_vin or "",
        "vehicle_year": repair_order.vehicle_year or "",
        "parts_total": parts_total,
        "labor_total": labor_total,
        "grand_total": parts_total + labor_total,
    }

    part_items = [
        {
            "article": m.matched_article or m.contract_article or "",
            "cat_number": m.nomenclature_cat_number or "",
            "name": m.matched_name or m.contract_name or "",
            "manufacturer": m.nomenclature_manufacturer or "",
            "unit": m.nomenclature_unit or "",
            "qty": float(m.contract_qty) if m.contract_qty is not None else 1.0,
            "price": float(m.matched_price) if m.matched_price is not None else "",
            "total": (
                (float(m.matched_price) if m.matched_price is not None else 0.0)
                * (float(m.contract_qty) if m.contract_qty is not None else 1.0)
            ),
            "warehouse": m.nomenclature_warehouse or "",
        }
        for m in part_matches
    ]
    labor_items = [
        {
            "description": l.matched_operation_name or l.description,
            "norm_hours": float(l.norm_hours) if l.norm_hours is not None else "",
            "hourly_rate": float(l.hourly_rate) if l.hourly_rate is not None else "",
            "total": float(l.total_cost) if l.total_cost is not None else "",
        }
        for l in labor_lines
    ]
    return context, part_items, labor_items


def generate_repair_order_document_from_template(
    repair_order: RepairOrder, template: DocumentTemplate
) -> tuple[str, list[str]]:
    context, part_items, labor_items = build_template_context(repair_order)
    output_dir = os.path.dirname(repair_order.storage_path)
    output_path = os.path.join(output_dir, f"repair_order_{repair_order.id}_final_{template.id}.xlsx")
    return document_template_engine.render_template(
        template.storage_path, output_path, context, part_items, labor_items
    )
